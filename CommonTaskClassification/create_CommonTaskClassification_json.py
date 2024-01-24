import pandas as pd
from openpyxl import load_workbook
import os
import json

# json文件需要填写的模板字典
template_dict_2 = {
    "commontask":[],
    "sub_commontask":[],
    # "feishu_document_link":[],
    "dataset_name_list":[],
    # 可以根据需要继续添加字段
}


# 读取Excel文件
# 获取当前脚本的路径
current_script_path = os.path.abspath(__file__)
# 获取当前脚本所在目录的路径
current_dir = os.path.dirname(current_script_path)
# 获取外层目录的路径
parent_dir = os.path.dirname(current_dir)
excel_file = os.path.join(parent_dir, 'data/数据接入文档信息汇总表.xlsx')
wb = load_workbook(excel_file)
sheet = wb.active


# 获取工作表
sheet_link = wb['Sheet1']
CommonTaskClassification_lists = ['begin']
grouped_lists = {}
for row_index, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
    if row[0] is not None and row[0] !='':
        if row[15] is not None and row[15] !='':
            print(row_index,row[1],row[15],row[5])
            
            sub_commontask_path = '/media/3_sdc/datasets/StandardDatasetTemplateFiles/CommonTaskClassification/'+row[15]+'/'+row[1]
            if sub_commontask_path in grouped_lists:
                grouped_lists[sub_commontask_path].append(row[4]+'/'+row[5])
            else:
                grouped_lists[sub_commontask_path] = [row[4]+'/'+row[5]]

for key, value_list in grouped_lists.items():
    # print(f"第1列值为 {key} 的所有第5列数据: {value_list}")

    output_dict = template_dict_2.copy()
    output_dict["commontask"] = key.split('/')[-2]
    output_dict["sub_commontask"] = key.split('/')[-1]
    # output_dict["feishu_document_link"] = md_content
    output_dict["dataset_name_list"] = list(set(value_list))
    json_filename = f'{key}/CommonTaskFields.json'

    with open(json_filename, 'w', encoding='utf-8') as jsonfile:
        json.dump(output_dict, jsonfile, ensure_ascii=False, indent=4)

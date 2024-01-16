import pandas as pd
from openpyxl import load_workbook
import os

def generate_table_rows(data_rows):
    table_rows = ""
    for row in data_rows:
        table_rows += """
    <tr> 
        <th> {categories} </th>  
        <th> {file_numbers} </th> 
        <th> {overall_description} </th> 
        <th> {segmentation_scenes} </th> 
        <th> {annotate_files} </th>  
        <th> {number_of_pictures} </th> 
        <th> {number_of_detection_frames} </th> 
        <th> {segment_description} </th>   
    </tr>""".format(**row)
    return table_rows

# 将生成的表格行插入到模板中
def create_table(template, header_data, row_data):
    table_rows = generate_table_rows(row_data)
    return template.format(**header_data, table_rows=table_rows)



# 在这里使用真实的数据替换占位符
filled_template = """
## {dataset_type}
飞书文档链接：{feishu_document_link}

创建目的: {creation_purpose}

数据存储规范: {data_storage_specifications}

## 数据集介绍

<table>
    <tr>
        <th rowspan="2"> 类别 </th> 
        <th rowspan="2"> 图片数量 </th> 
        <th rowspan="2"> 整体描述 </th> 
        <th colspan="5"> 划分(split) </th>  
    </tr>
    <tr> 
        <td> 细分场景 </td>
        <td> 标注文件 </td>
        <td> 图片数量 </td>
        <td> 检测框数量 </td>
        <td> 细分描述 </td>
    </tr>{table_rows}
</table>
"""


# 读取Excel文件
excel_file = 'data/数据接入文档信息汇总表.xlsx'
wb = load_workbook(excel_file)
sheet = wb.active



# 获取工作表
sheet_link = wb['Sheet1']

# # 获取某个单元格
# cell = sheet_link['A3'] 

# # 检查单元格是否有超链接
# if cell.hyperlink:
#     print("这个单元格有超链接:", cell.hyperlink.target)
# else:
#     print("这个单元格没有超链接")

# # 提取表头
# headers = []
# for row in sheet.iter_rows(min_row=1, max_row=2, values_only=True):
#     print(row)
#     for header in row:
#         headers.append(header)

# # 判断最大列数
# max_column = sheet.max_column
# #print(max_column)

# 提取数据并建立Markdown文件
dataset_name = ['begin']
row_data = []
header_data_lists = [] 
dataset_roor_list = []

j = 0
for row_index, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
    #print(row_index, row[4], row[5])
    
    # # 检查文件夹路径是否存在
    # dataset_path = row[4]+ '/'+row[5]
    # is_dir = os.path.isdir(dataset_path)
    # if is_dir:
    #     continue
    # else:
    #     print(f"{row_index, dataset_path,row[5], row[6]} 不存在,保存README.md失败")

    cell_num_str = 'A'+str(row_index)
    # cell = sheet_link['A3'] 
    cell = sheet_link[cell_num_str] 
    if cell.hyperlink:
        # 提取单元格文本和链接
        text = cell.value
        link = cell.hyperlink.target
        
        # 将信息写成Markdown链接
        md_content = f'[{text}]({link})  \n'
    else:
        print('单元格无超链接')
    

    if row[6]:
        # print(row_index, row)
        #print(row[5])
        dataset_name.append(row[5])  
        dataset_roor_list.append(row[4])
        
        header_data = {
            'dataset_type': row[1],
            'creation_purpose': row[2],
            'data_storage_specifications': row[3],
            'feishu_document_link':md_content,
        }  
        header_data_lists.append(header_data)  
        
        value_data = {
            'categories': row[7],
            'file_numbers': row[8],
            'overall_description': row[9],
            'segmentation_scenes': row[10],
            'annotate_files': row[11],
            'number_of_pictures': row[12],
            'number_of_detection_frames': row[13],
            'segment_description': row[14],
        }
        #print(value_data)
        
        if row_index==3:
            #print("第一行数据")
            row_data.append(value_data)

        else:
            if row[5]==dataset_name[-2]:
                row_data.append(value_data)
                #print(row_index, row)

                # 为了保存最后一行数据而添加的代码
                final_html = create_table(filled_template, header_data_lists[-1], row_data)
                # print(row[5], final_html)
                # readme_filename = f"output/README_{row[5]}.md"
                dataset_path = dataset_roor_list[-2] + '/' + row[5]
                readme_filename = dataset_path + "/README.md"
                with open(readme_filename, "w", encoding="utf-8") as md_file:
                    md_file.write(final_html)
                print(f"Markdown文件已生成: {readme_filename}")

            else:
                #print(row_data)
                final_html = create_table(filled_template, header_data_lists[-2], row_data)
                # print(row[5], final_html)
                # readme_filename = f"output/README_{dataset_name[-2]}.md"
                dataset_path = dataset_roor_list[-2] + '/' + dataset_name[-2]
                readme_filename = dataset_path + "/README.md"
                with open(readme_filename, "w", encoding="utf-8") as md_file:
                    md_file.write(final_html)
                print(f"Markdown文件已生成: {readme_filename}")
                
                row_data = []    
                row_data.append(value_data)
                
                

                # 为了保存最后一行数据而添加的代码
                final_html = create_table(filled_template, header_data_lists[-1], row_data)
                # print(row[5], final_html)
                # readme_filename = f"output/README_{row[5]}.md"
                dataset_path = dataset_roor_list[-1] + '/' + row[5]
                readme_filename = dataset_path + "/README.md"
                with open(readme_filename, "w", encoding="utf-8") as md_file:
                    md_file.write(final_html)
                    j = j+1
                print(f"Markdown文件已生成: {readme_filename}")
print(j)


